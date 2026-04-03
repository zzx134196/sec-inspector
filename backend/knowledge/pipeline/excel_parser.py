"""Excel测评模板解析器 — 将等保测评参考描述Excel解析为结构化数据"""
import os
import re
from typing import List, Dict, Any, Optional

from loguru import logger


def parse_excel_template(file_path: str) -> List[Dict[str, Any]]:
    """
    解析等保测评参考描述Excel文件
    :param file_path: Excel文件路径
    :return: 结构化模板记录列表
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")

    try:
        import openpyxl
    except ImportError:
        raise ImportError("需要安装openpyxl: pip install openpyxl")

    file_name = os.path.basename(file_path)
    category = _infer_category(file_name)

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    all_records = []

    for sheet_name in wb.sheetnames:
        # 跳过文档信息表等非数据sheet
        if _is_meta_sheet(sheet_name):
            logger.debug(f"跳过元数据sheet: {sheet_name}")
            continue

        ws = wb[sheet_name]
        object_type = _infer_object_type(sheet_name, file_name)

        try:
            records = _parse_sheet(ws, sheet_name, category, object_type, file_name)
            all_records.extend(records)
            logger.info(f"解析sheet '{sheet_name}' 完成，获取 {len(records)} 条记录")
        except Exception as e:
            logger.warning(f"解析sheet '{sheet_name}' 失败: {e}")
            continue

    wb.close()
    logger.info(f"Excel解析完成: {file_name}, 共 {len(all_records)} 条记录")
    return all_records


def _parse_sheet(
    ws,
    sheet_name: str,
    category: str,
    object_type: str,
    source_file: str,
) -> List[Dict[str, Any]]:
    """解析单个sheet为记录列表"""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    # 查找表头行
    header_row_idx, col_map = _find_header(rows)
    if header_row_idx is None:
        logger.debug(f"Sheet '{sheet_name}' 未找到有效表头")
        return []

    records = []
    current_control_point = ""
    item_index = 0

    for row_idx in range(header_row_idx + 1, len(rows)):
        row = rows[row_idx]
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        # 提取字段
        record = _extract_fields(row, col_map)
        if not record:
            continue

        # 更新控制点（如果本行有新的控制点）
        cp = record.get("control_point", "").strip()
        if cp:
            current_control_point = cp
        else:
            record["control_point"] = current_control_point

        # 跳过无实质内容的行
        has_content = any([
            record.get("control_item"),
            record.get("compliant_desc"),
            record.get("non_compliant_desc"),
        ])
        if not has_content:
            continue

        item_index += 1
        record["category"] = category
        record["object_type"] = object_type
        record["source_file"] = source_file
        record["item_index"] = item_index

        records.append(record)

    return records


def _find_header(rows: list) -> tuple:
    """在前10行中查找表头行，返回 (行索引, 列名→列索引映射)"""
    # 已知的列名关键词及其标准化映射
    KNOWN_HEADERS = {
        "控制点": "control_point",
        "安全控制点": "control_point",
        "控制项": "control_item",
        "测评项": "control_item",
        "符合": "compliant_desc",
        "部分符合": "partial_compliant_desc",
        "不符合": "non_compliant_desc",
        "不适用": "not_applicable_desc",
        "高风险": "high_risk_criteria",
        "高风险判例": "high_risk_criteria",
        "高风险判定": "high_risk_criteria",
        "问题描述": "problem_desc",
        "问题分析": "problem_analysis",
        "危害分析": "harm_analysis",
        "整改建议": "fix_suggestion",
        "备注": "remarks",
        "测评方法": "remarks",
        "序号": "seq_no",
        "编号": "test_item_number",
        "测评项编号": "test_item_number",
        "ObjectType": "object_type_col",
        "TestItemNumber": "test_item_number",
        "StdCode": "std_code",
        "降风险": "risk_reduction",
    }

    scan_limit = min(10, len(rows))
    for i in range(scan_limit):
        row = rows[i]
        if not row:
            continue

        col_map = {}
        match_count = 0

        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            cell_str = str(cell).strip()
            for keyword, field_name in KNOWN_HEADERS.items():
                if keyword in cell_str:
                    col_map[field_name] = col_idx
                    match_count += 1
                    break

        # 至少匹配到2个已知列名才认为是表头
        if match_count >= 2:
            return i, col_map

    return None, {}


def _extract_fields(row: tuple, col_map: dict) -> Optional[Dict[str, Any]]:
    """从一行数据中提取字段"""
    if not col_map:
        return None

    def get_val(field_name: str) -> str:
        idx = col_map.get(field_name)
        if idx is not None and idx < len(row) and row[idx] is not None:
            return str(row[idx]).strip()
        return ""

    record = {
        "control_point": get_val("control_point"),
        "control_item": get_val("control_item"),
        "test_item_number": get_val("test_item_number"),
        "std_code": get_val("std_code"),
        "compliant_desc": get_val("compliant_desc"),
        "partial_compliant_desc": get_val("partial_compliant_desc"),
        "non_compliant_desc": get_val("non_compliant_desc"),
        "not_applicable_desc": get_val("not_applicable_desc"),
        "problem_desc": get_val("problem_desc"),
        "problem_analysis": get_val("problem_analysis"),
        "harm_analysis": get_val("harm_analysis"),
        "fix_suggestion": get_val("fix_suggestion"),
        "high_risk_criteria": get_val("high_risk_criteria"),
        "risk_reduction": get_val("risk_reduction"),
        "remarks": get_val("remarks"),
    }

    return record


def _infer_category(file_name: str) -> str:
    """从文件名推断安全层面分类"""
    mapping = {
        "安全计算环境": "安全计算环境",
        "安全通信网络": "安全通信网络",
        "安全区域边界": "安全区域边界",
        "安全物理环境": "安全物理环境",
        "安全管理中心": "安全管理中心",
        "安全管理制度": "安全管理制度",
        "安全管理机构": "安全管理机构",
        "安全管理人员": "安全管理人员",
        "安全建设管理": "安全建设管理",
        "安全运维管理": "安全运维管理",
        "管理类": "安全管理",
        "高风险判定": "高风险判定指引",
    }
    for keyword, cat in mapping.items():
        if keyword in file_name:
            return cat
    return "其他"


def _infer_object_type(sheet_name: str, file_name: str) -> str:
    """从sheet名称和文件名推断测评对象类型"""
    # 先看sheet名称
    known_objects = [
        "Linux", "Windows", "Android", "iOS",
        "MySQL", "Oracle", "SQL Server", "PostgreSQL", "Redis", "MongoDB",
        "Tomcat", "WebSphere", "IIS", "Nginx", "WebLogic", "Apache",
        "VMware", "Docker", "Kubernetes",
        "华为", "H3C", "Cisco", "锐捷", "深信服",
    ]
    for obj in known_objects:
        if obj.lower() in sheet_name.lower():
            return obj

    # sheet名称本身可能就是对象类型
    clean_name = sheet_name.strip()
    if clean_name and clean_name not in ("Sheet1", "Sheet2", "Sheet3", "文档信息表"):
        return clean_name

    # 从文件名推断
    if "操作系统" in file_name:
        return "操作系统"
    if "数据库" in file_name:
        return "数据库"
    if "中间件" in file_name:
        return "中间件"
    if "网络设备" in file_name:
        return "网络设备"
    if "安全设备" in file_name:
        return "安全设备"

    return "通用"


def _is_meta_sheet(sheet_name: str) -> bool:
    """判断是否为元数据sheet（非数据内容）"""
    meta_keywords = ["文档信息", "修订记录", "版本", "说明", "目录", "封面"]
    for kw in meta_keywords:
        if kw in sheet_name:
            return True
    return False


def parse_high_risk_excel(file_path: str) -> List[Dict[str, Any]]:
    """
    专门解析高风险判定指引Excel
    :param file_path: 高风险判定指引Excel路径
    :return: 高风险判例记录列表
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")

    try:
        import openpyxl
    except ImportError:
        raise ImportError("需要安装openpyxl: pip install openpyxl")

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    records = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            continue

        # 查找表头
        header_row_idx = None
        headers = []
        for i in range(min(10, len(rows))):
            row = rows[i]
            if row and any(cell and ("高风险" in str(cell) or "安全层面" in str(cell) or "控制点" in str(cell)) for cell in row if cell):
                header_row_idx = i
                headers = [str(cell).strip() if cell else "" for cell in row]
                break

        if header_row_idx is None:
            continue

        for row_idx in range(header_row_idx + 1, len(rows)):
            row = rows[row_idx]
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            record = {}
            for col_idx, header in enumerate(headers):
                if col_idx < len(row) and row[col_idx] is not None:
                    record[header] = str(row[col_idx]).strip()

            if record:
                records.append(record)

    wb.close()
    logger.info(f"高风险判定指引解析完成: {len(records)} 条记录")
    return records
